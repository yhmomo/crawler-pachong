import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from Tool.Strclean import clean_string

# 读取Excel文件
# file_path = '本科普通批上.xlsx'
file_path = '本科普通批下.xlsx'
dfs = pd.read_excel(file_path, engine='openpyxl', sheet_name=None)

output_file_path = 'output_1.xlsx'

# 使用 ExcelWriter 将所有工作表的数据写入同一个文件
with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    for sheet_name, df in dfs.items():
        # 对DataFrame的每一列应用清理函数
        df_below_key = df.map(clean_string)
        # 删除含有关键字的所在行
        keywords = ['https', 'GA0', 'XUN', '第五部分', '下列院校', '招生计划', '前安排一个', '特别提醒',
                    '下列院校专业组(专业)志愿填报在本科普通批单设志愿栏', '湖北招生考试',
                    '具备高校专项计划报考资格的考生方可填报', '根据教育部要求', '已经报名', '考生可填报',
                    '院校办学性质和办学地点']
        for index, row in df_below_key.iterrows():
            for keyword in keywords:
                if row.astype(str).str.contains(keyword, regex=False).any():
                    df_below_key = df_below_key.drop(index)
                    break  # 如果找到关键字，删除该行并跳出内层循环

        # 删除全为空值的列
        df_below_key = df_below_key.dropna(axis=1, how='all')

        # 重置索引
        df_below_key = df_below_key.reset_index(drop=True)

        # 将清理后的数据写入新的Excel文件中的一个工作表
        df_below_key.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"已处理工作表：{sheet_name}")

# 加载生成的Excel文件
workbook = load_workbook(filename=output_file_path)

# 遍历所有工作表并自动调整列宽
for sheet_name in workbook.sheetnames:
    worksheet = workbook[sheet_name]
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                # Measure the length of the cell value
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width

# 保存调整后的Excel文件
workbook.save(output_file_path)
print(f"清理后的数据已保存到 {output_file_path}，并且列宽已自动调整")
